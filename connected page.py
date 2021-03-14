from tkinter import *
def LoginPage():
    
    
    login_screen=Tk()
    login_screen.title("Login")
    login_screen.geometry("300x250")
    
   
    Label(login_screen, text="Please enter login details").pack()
    Label(login_screen, text="").pack()
    Label(login_screen, text="Username").pack()
    username_login_entry = Entry(login_screen, textvariable="username")
    username_login_entry.pack()
    Label(login_screen, text="").pack()
    Label(login_screen, text="Password").pack()
    password__login_entry = Entry(login_screen, textvariable="password", show= '*')
    password__login_entry.pack()
    Label(login_screen, text="").pack()
    Button(login_screen, text="Login", width=10, height=1).pack()
    login_screen.mainloop()
LoginPage()
from tkinter import messagebox
from openpyxl import *
from tkinter  import ttk
from tkcalendar import DateEntry
import random
#opening that excel sheet
wb=load_workbook('C:\\Users\\Aniket Singh\\OneDrive\\Desktop\\singupdata.xlsx')
sheet=wb.active
def excel():
    
    
    #set the heading to excel sheet.
    sheet.column_dimensions['A'].width=50
    sheet.cell(row=1,column=1).value="name"
    
    sheet.column_dimensions['B'].width=50
    sheet.cell(row=1,column=2).value="fathername"
    
    sheet.column_dimensions['C'].width=50
    sheet.cell(row=1,column=3).value="mothername"
    
    sheet.column_dimensions['D'].width=50
    sheet.cell(row=1,column=4).value="gender"
    
    sheet.column_dimensions['E'].width=50
    sheet.cell(row=1,column=5).value="state"
   
    sheet.column_dimensions['f'].width=50
    sheet.cell(row=1,column=6).value="DOB"
    
    sheet.column_dimensions['G'].width=50
    sheet.cell(row=1,column=7).value="email"
    
    sheet.column_dimensions['F'].width=50
    sheet.cell(row=1,column=8).value="password"
def insert_data():
    #featch the data from interface
    user_name=name_data.get()
    father_name=name_data1.get()
    mother_name=name_data2.get()
    gender=var.get()
    state=cm.get()
    DOB=a.get()
    email=name_data3.get()
    pass1=password1.get()
    
    if(user_name ==" " or father_name ==" " or mother_name ==" " or email ==" " or password1==" "):
        messagebox.showerror("error","enter com plete detail")
    else:
        
        #print (user_name)
        #insert the data
        current_row=sheet.max_row
        sheet.cell(row=current_row+1,column=1).value=user_name
        sheet.cell(row=current_row+1,column=2).value=father_name
        sheet.cell(row=current_row+1,column=3).value=mother_name
        sheet.cell(row=current_row+1,column=7).value=email
        sheet.cell(row=current_row+1,column=4).value=gender
        sheet.cell(row=current_row+1,column=5).value=state
        sheet.cell(row=current_row+1,column=6).value=DOB
        sheet.cell(row=current_row+1,column=8).value=pass1
    
    
        #save the data
        wb.save('C:\\Users\\Aniket Singh\\OneDrive\\Desktop\\singupdata.xlsx')
        #giving alert to user
        # there is 3 diffrent type of alert
        
       # messagebox.showerror("error","this is an error message")
       
        #messagebox.showwarnin("warning","this is the warning")
        
        messagebox.showinfo("information","conformation message")
w=Tk()
w.geometry('600x800')
w.configure(bg='green')

heading=Label(w,text="registration")
heading.place(x=200,y=30)

name=Label(w,text="enter name")
name.place(x=20,y=60)
name_data=Entry(w)
name_data.place(x=130,y=60)

fathername=Label(w,text="enter fathername")
fathername.place(x=20,y=90)
name_data1=Entry(w)
name_data1.place(x=130,y=90)

mothername=Label(w,text="enter mothername")
mothername.place(x=20,y=120)
name_data2=Entry(w)
name_data2.place(x=130,y=120)

#email
email=Label(w,text="email")
email.place(x=20,y=240)
name_data3=Entry(w)
name_data3.place(x=130,y=240)

gender=Label(w,text="gender")
gender.place(x=20,y=150)
#featch the data from radiobutton
var=StringVar()#this function cereat variable which conatin an sting value
var.set(FALSE)
#var1=IntVar()
male_rd=Radiobutton(w,text="male",variable=var,value="male")
male_rd.place(x=120,y=150)

female_rd=Radiobutton(w,text="female",variable=var,value="female")
female_rd.place(x=210,y=150)
#read a data from combobox
cm =ttk.Combobox(w)
cm ['values']=('bihar','up','delhi')
cm.current()
cm.place(x=20,y=180)
#dob
a=DateEntry(w,width=12,year=2000,month=5,day=23,bg='blue')
a.place(x=20,y=210)

password=Label(w,text="enter the password")
password.place(x=20,y=270)
password1=Entry(w,text="password",show="*")
password1.place(x=130,y=270)

b=Button(w,text="singup",command=insert_data)
b.place(x=20,y=300)
excel()

b1=Button(w,text="login",command=LoginPage)
b1.place(x=500,y=300)






w.mainloop()
